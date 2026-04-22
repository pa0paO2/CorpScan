"""
结果导出模块 - 支持 CSV 和 Excel 格式

使用：
    from utils.exporter import export_results
    export_results(results, format="csv", filename="output/result.csv")
"""

import csv
import json
from pathlib import Path
from datetime import datetime
from typing import List, Optional
from models import AssetType


def ensure_output_dir(filename: str) -> Path:
    """确保输出目录存在"""
    filepath = Path(filename)
    output_dir = filepath.parent
    if output_dir and not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)
    return filepath


def flatten_dict(d: dict, parent_key: str = "", sep: str = "_") -> dict:
    """将嵌套字典扁平化（用于 CSV 导出）"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


def model_to_row(item) -> dict:
    """将 CompanyModel 转换为字典（优化字段输出）"""
    row = {
        "公司名称": item.name or "-",
        "资产类型": item.asset_type or "-",
        "域名": item.domain or "-",
        "ICP备案号": item.icp_number or "-",
        "网站名称": item.site_name or "-",
        "APP名称": item.app_name or "-",
        "APP包名": item.app_package or "-",
        "小程序名称": item.miniapp_name or "-",
        "小程序ID": item.miniapp_id or "-",
        "法定代表人": item.legal_person or "-",
        "经营状态": item.status or "-",
    }

    # 添加子公司关联信息（如果存在）
    if item.extra:
        if "parentCompany" in item.extra:
            row["母公司"] = item.extra.get("parentCompany", "-")
        if "subsidiaryName" in item.extra:
            row["子公司"] = item.extra.get("subsidiaryName", "-")
        if "equityPercent" in item.extra:
            row["股权比例"] = item.extra.get("equityPercent", "-")

        # 其他关键扩展字段
        for key in ["equityRatio", "registerStatus", "industry", "companyType", "investAmount",
                    "appType", "classes", "examineDate"]:
            if key in item.extra:
                row[key] = item.extra.get(key, "-")

    return row


def export_to_csv(results: List, filename: str) -> str:
    """
    导出结果到 CSV 文件

    :param results: CompanyModel 列表
    :param filename: 输出文件名
    :return: 实际保存的文件路径
    """
    if not results:
        return ""

    filepath = ensure_output_dir(filename)

    # 收集所有可能的列（处理动态字段）
    all_columns = set()
    rows_data = []

    for item in results:
        row = model_to_row(item)
        all_columns.update(row.keys())
        rows_data.append(row)

    # 固定列顺序（优化后的字段）
    fixed_cols = [
        "公司名称", "资产类型", "母公司", "子公司", "股权比例",
        "域名", "ICP备案号", "网站名称",
        "APP名称", "APP包名",
        "小程序名称", "小程序ID",
        "法定代表人", "经营状态",
        "注册状态", "行业", "公司类型", "投资金额",
        "appType", "classes", "examineDate"
    ]

    # 其他字段排序后追加
    extra_cols = sorted([c for c in all_columns if c not in fixed_cols])
    columns = fixed_cols + extra_cols

    # 写入 CSV
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()
        for row in rows_data:
            writer.writerow(row)

    return str(filepath.absolute())


def export_to_excel(results: List, filename: str) -> str:
    """
    导出结果到 Excel 文件（多工作表，按资产类型分组）

    :param results: CompanyModel 列表
    :param filename: 输出文件名
    :return: 实际保存的文件路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        # 如果没有 openpyxl，回退到 CSV
        csv_file = filename.replace(".xlsx", ".csv")
        return export_to_csv(results, csv_file)

    if not results:
        return ""

    filepath = ensure_output_dir(filename)

    # 创建工作簿
    wb = Workbook()

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 按资产类型分组
    grouped = {}
    for item in results:
        t = item.asset_type or "其他"
        if t not in grouped:
            grouped[t] = []
        grouped[t].append(item)

    # 删除默认工作表
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 为每种资产类型创建工作表
    for idx, (asset_type, items) in enumerate(grouped.items()):
        ws = wb.create_sheet(title=asset_type[:31])  # Excel 工作表名最多 31 字符

        # 辅助函数：获取子公司信息
        def get_sub_info(item):
            parent = item.extra.get("parentCompany", "") if item.extra else ""
            sub = item.extra.get("subsidiaryName", "") if item.extra else ""
            equity = item.extra.get("equityPercent", "") if item.extra else ""
            return parent, sub, equity

        # 辅助函数：获取审核日期
        def get_examine_date(item):
            if item.extra:
                return item.extra.get("examineDate", "-") or "-"
            return "-"

        # 根据类型确定列
        if asset_type == AssetType.WEBSITE.value:
            headers = ["序号", "公司名称", "网站名称", "域名", "ICP备案号", "审核日期", "母公司", "子公司", "股权比例"]
            def get_row_data(i, item):
                parent, sub, equity = get_sub_info(item)
                return [i, item.name, item.site_name, item.domain, item.icp_number,
                        get_examine_date(item), parent, sub, equity]
        elif asset_type == AssetType.APP.value:
            headers = ["序号", "公司名称", "APP名称", "类型", "分类", "审核日期", "母公司", "子公司", "股权比例"]
            def get_row_data(i, item):
                parent, sub, equity = get_sub_info(item)
                return [i, item.name, item.app_name,
                        item.extra.get("appType", "应用") if item.extra else "应用",
                        item.extra.get("classes", "-") if item.extra else "-",
                        get_examine_date(item), parent, sub, equity]
        elif asset_type == AssetType.MINIPROGRAM.value:
            headers = ["序号", "公司名称", "小程序名称", "备案号", "审核日期", "母公司", "子公司", "股权比例"]
            def get_row_data(i, item):
                parent, sub, equity = get_sub_info(item)
                return [i, item.name, item.miniapp_name, item.icp_number,
                        get_examine_date(item), parent, sub, equity]
        elif asset_type == AssetType.SUBSIDIARY.value:
            headers = ["序号", "公司名称", "股权比例", "注册状态", "行业", "母公司"]
            def get_row_data(i, item):
                parent = item.extra.get("parentName", "") if item.extra else ""
                return [i, item.name,
                        item.extra.get("equityPercent", "-") if item.extra else "-",
                        item.extra.get("registerStatus", "-") if item.extra else "-",
                        item.extra.get("industry", "-") if item.extra else "-",
                        parent]
        else:
            headers = ["序号", "公司名称", "资产类型", "域名", "备案号"]
            def get_row_data(i, item):
                return [i, item.name, item.asset_type, item.domain, item.icp_number]

        # 写入表头
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for i, item in enumerate(items, 1):
            row_data = get_row_data(i, item)
            ws.append(row_data)

        # 设置列宽和样式
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell.border = thin_border
                    cell.alignment = cell_alignment
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # 冻结首行
        ws.freeze_panes = "A2"

    # 保存文件
    wb.save(filepath)
    return str(filepath.absolute())


def export_results(results: List, format: str = "csv", filename: Optional[str] = None) -> str:
    """
    导出结果到文件（统一入口）

    :param results: CompanyModel 列表
    :param format: 导出格式（csv/excel/xlsx）
    :param filename: 输出文件名（可选，默认自动生成）
    :return: 实际保存的文件路径
    """
    # 生成默认文件名
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ext = "xlsx" if format in ["excel", "xlsx"] else "csv"
        filename = f"output/result_{timestamp}.{ext}"

    # 根据格式选择导出方式
    if format in ["excel", "xlsx"]:
        filepath = export_to_excel(results, filename)
        if filepath.endswith(".csv"):
            print(f"[!] 未安装 openpyxl，已回退导出为 CSV: {filepath}")
        else:
            print(f"[+] 已导出 Excel: {filepath}")
    else:
        filepath = export_to_csv(results, filename)
        print(f"[+] 已导出 CSV: {filepath}")

    return filepath


if __name__ == "__main__":
    # 测试导出功能
    from models import CompanyModel

    test_data = [
        CompanyModel(
            name="小米科技",
            source="tianyancha",
            asset_type=AssetType.WEBSITE.value,
            site_name="小米商城",
            domain="mi.com",
            icp_number="京ICP备10046444号",
            extra={"examineDate": "2024-01-01"}
        ),
        CompanyModel(
            name="小米科技",
            source="tianyancha",
            asset_type=AssetType.APP.value,
            app_name="小米商城",
            extra={"appType": "应用", "classes": "电商"}
        ),
    ]

    export_results(test_data, "csv", "output/test.csv")
    export_results(test_data, "excel", "output/test.xlsx")
